import openpyxl as oxl
from openpyxl.chart import BarChart, Reference

def process_workbook():
    wb=oxl.load_workbook('excel_eg.xlsx')
    sheet=wb['Sheet1']

    for row in range(2,sheet.max_row+1):
        price_cell=sheet.cell(row,3)
        corrected_price=price_cell.value*.9         #decrease the price by 10%
        corrected_price_cell=sheet.cell(row,4)
        corrected_price_cell.value=corrected_price

    corrected_price_cell=sheet['d1']
    corrected_price_cell.value='corrected price'

    values=Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')

    wb.save('excel_eg2.xlsx')

process_workbook()
